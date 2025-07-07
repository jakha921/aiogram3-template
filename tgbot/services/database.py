import aiogram
from typing import AsyncGenerator

import aiomysql
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

from tgbot.config import Config
from tgbot.services.db_base import Base
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import tempfile
import os


async def create_db_session(config: Config) -> AsyncGenerator[AsyncSession, None]:
    """Create DB session and handle database creation if needed"""

    # Database connection parameters
    auth_data = {
        "user": config.db.user,
        "password": config.db.password,
        "host": config.db.host,
        "port": config.db.port,
        "database": config.db.database,
    }

    print('mysql+aiomysql://{user}:{password}@{host}:{port}/{database}'.format(**auth_data))

    # Check if database exists and create if needed
    try:
        # Подключаемся без выбора БД
        conn = await aiomysql.connect(
            host=auth_data["host"],
            port=int(auth_data["port"]),
            user=auth_data["user"],
            password=auth_data["password"]
        )
        async with conn.cursor() as cur:
            await cur.execute(f"CREATE DATABASE IF NOT EXISTS `{config.db.database}`")
        conn.close()
    except Exception as e:
        print(f"Database initialization error: {e}")
        raise

    # Create async engine
    engine = create_async_engine(
        f"mysql+aiomysql://{auth_data['user']}:{auth_data['password']}@"
        f"{auth_data['host']}:{auth_data['port']}/{auth_data['database']}",
        echo=False,
        future=True,
        pool_pre_ping=False
    )

    # Create tables
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    # Create session factory
    async_session = sessionmaker(
        engine,
        class_=AsyncSession,
        expire_on_commit=False
    )

    return async_session

def generate_reconciliation_act_excel(summary, customer_name, period_str, company_name="AVTOLIDER"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Акт сверки"

    # Заголовок
    ws.merge_cells('A1:F1')
    ws['A1'] = "Акт сверки"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:F2')
    ws['A2'] = f"за период: {period_str}"
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A3:F3')
    ws['A3'] = f"между: {company_name} и {customer_name}"
    ws['A3'].alignment = Alignment(horizontal="center")

    ws.append(["Дата", "Документ", "Сумма", "Оплачено", "Долг", "Примечание"])
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in summary:
        ws.append([
            row.get("Дата", ""),
            row.get("Документ", ""),
            row.get("Сумма", 0),
            row.get("Оплачено", 0),
            row.get("Долг", 0),
            row.get("Примечание", "")
        ])

    # Итоговая строка
    last_row = ws.max_row + 1
    ws[f'A{last_row}'] = "Итого долг:"
    ws[f'E{last_row}'] = sum(float(r.get("Долг", 0) or 0) for r in summary)
    ws[f'A{last_row}'].font = Font(bold=True)
    ws[f'E{last_row}'].font = Font(bold=True)

    # Стили
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Сохраняем во временный файл
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
