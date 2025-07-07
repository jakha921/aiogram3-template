import os
from datetime import datetime
from typing import List, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from num2words import num2words


async def generate_invoice_excel(
    invoice_data: List[Tuple],
    invoice_number: str = None,
    output_dir: str = "invoices"
) -> str:
    """
    Генерирует накладную в Excel формате из переданных данных.
    
    Args:
        invoice_data: Список кортежей с данными товаров
        invoice_number: Номер накладной (если не указан, генерируется автоматически)
        output_dir: Директория для сохранения файла
        
    Returns:
        Путь к созданному Excel файлу
    """
    # Создаем директорию если не существует
    os.makedirs(output_dir, exist_ok=True)
    
    # Генерируем номер накладной если не указан
    if not invoice_number:
        invoice_number = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # Создаем новый workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Накладная"
    
    # Настройка стилей
    header_font = Font(name='Arial', size=16, bold=True)
    title_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # Заголовок накладной
    ws.merge_cells('A1:I2')
    ws['A1'] = f'ТОВАРНАЯ НАКЛАДНАЯ № {invoice_number}'
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    
    # Дата накладной
    ws.merge_cells('A3:I3')
    ws['A3'] = f'от {datetime.now().strftime("%d.%m.%Y")}'
    ws['A3'].font = normal_font
    ws['A3'].alignment = center_alignment
    
    # Информация о компании (из первой записи)
    if invoice_data:
        company_name = invoice_data[0][0]
        ws.merge_cells('A5:D5')
        ws['A5'] = f'Поставщик: {company_name}'
        ws['A5'].font = bold_font
        ws['A5'].alignment = left_alignment
    
    # Заголовки таблицы
    headers = ['№', 'Код товара', 'Наименование', 'Количество', 'Ед.изм.', 
               'Цена', 'Сумма', 'Статус', 'Дата']
    
    start_row = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = title_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Заполнение данными
    total_sum = 0
    current_row = start_row + 1
    
    for idx, item in enumerate(invoice_data, 1):
        # Распаковка данных
        company, code, name, date, operation, quantity, price, amount, status = item
        
        # Заполнение строки
        ws.cell(row=current_row, column=1, value=idx).alignment = center_alignment
        ws.cell(row=current_row, column=2, value=code).alignment = center_alignment
        ws.cell(row=current_row, column=3, value=name).alignment = left_alignment
        ws.cell(row=current_row, column=4, value=quantity).alignment = center_alignment
        ws.cell(row=current_row, column=5, value='шт.').alignment = center_alignment
        ws.cell(row=current_row, column=6, value=f"{price:,.2f}").alignment = right_alignment
        ws.cell(row=current_row, column=7, value=f"{amount:,.2f}").alignment = right_alignment
        ws.cell(row=current_row, column=8, value=status).alignment = center_alignment
        ws.cell(row=current_row, column=9, value=date.strftime("%d.%m.%Y %H:%M")).alignment = center_alignment
        
        # Применение границ
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).font = normal_font
        
        total_sum += amount
        current_row += 1
    
    # Итоговая строка
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = 'ИТОГО:'
    ws[f'A{current_row}'].font = bold_font
    ws[f'A{current_row}'].alignment = right_alignment
    ws[f'A{current_row}'].border = thin_border
    
    ws[f'G{current_row}'] = f"{total_sum:,.2f}"
    ws[f'G{current_row}'].font = bold_font
    ws[f'G{current_row}'].alignment = right_alignment
    ws[f'G{current_row}'].border = thin_border
    
    # Подписи
    current_row += 3
    ws[f'A{current_row}'] = 'Отпустил: ________________'
    ws[f'F{current_row}'] = 'Получил: ________________'
    
    current_row += 2
    ws[f'A{current_row}'] = 'М.П.'
    ws[f'F{current_row}'] = 'М.П.'
    
    # Настройка ширины колонок
    column_widths = [5, 12, 40, 12, 8, 15, 15, 20, 18]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Сохранение файла
    filename = f"invoice_{invoice_number}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filepath


async def generate_reconciliation_act_excel(
    act_data: list,
    company1: str,
    company2: str,
    period_start: str,
    period_end: str,
    saldo_start: float,
    saldo_end: float,
    output_dir: str = "invoices"
) -> str:
    """
    Генерирует акт сверки в Excel формате по примеру с фото.
    Args:
        act_data: список словарей с данными по операциям (выход get_customer_sales_summary)
        company1: название первой стороны (например, AVTOLIDER)
        company2: название второй стороны
        period_start: дата начала периода (строка)
        period_end: дата конца периода (строка)
        saldo_start: начальное сальдо
        saldo_end: конечное сальдо
        output_dir: папка для сохранения
    Returns:
        Путь к созданному Excel-файлу
    """
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Акт сверки"

    # Стили
    bold = Font(bold=True, size=12)
    big_bold = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовок
    ws.merge_cells('C3:F3')
    ws['C3'] = "Акт сверки"
    ws['C3'].font = big_bold
    ws['C3'].alignment = center

    ws.merge_cells('C4:F4')
    ws['C4'] = f"взаимных расчетов за период: {period_start} - {period_end}"
    ws['C4'].alignment = center
    ws.merge_cells('C5:F5')
    ws['C5'] = f"между: {company1} и {company2}"
    ws['C5'].alignment = center

    ws.merge_cells('B7:G7')
    ws['B7'] = f"Мы, нижеподписавшиеся, {company1} с одной стороны, и {company2}, с другой стороны,\nсоставили данный акт сверки в том, что, состояние взаимных расчетов по данным учета следующее:"
    ws['B7'].alignment = Alignment(wrap_text=True)

    # Таблица
    ws['B9'] = "Дата"
    ws['C9'] = "Документ"
    ws['D9'] = f"{company1}"
    ws['E9'] = "Дебет"
    ws['F9'] = "долг"
    ws['G9'] = f"{company2}"
    ws['H9'] = "Дебет"
    ws['I9'] = "Кредит"
    for col in range(2, 10):
        ws.cell(row=9, column=col).font = bold
        ws.cell(row=9, column=col).alignment = center
        ws.cell(row=9, column=col).border = border

    # Сальдо начальное
    ws.merge_cells('B10:C10')
    ws['B10'] = "Сальдо начальное"
    ws['D10'] = f"{saldo_start:,.2f}"
    ws['E10'] = "0.00"
    ws['F10'] = "0.00"
    ws['G10'] = "0.00"
    ws['H10'] = "0.00"
    ws['I10'] = "0.00"
    for col in range(2, 10):
        ws.cell(row=10, column=col).alignment = center
        ws.cell(row=10, column=col).border = border

    # Операции
    row = 11
    for op in act_data:
        ws[f'B{row}'] = op['Дата'].strftime('%-m/%-d/%Y') if hasattr(op['Дата'], 'strftime') else str(op['Дата'])
        ws[f'C{row}'] = op['Документ']
        ws[f'D{row}'] = f"{op['Сумма']:,.2f}"
        ws[f'E{row}'] = "0.00"
        ws[f'F{row}'] = "0.00"
        ws[f'G{row}'] = f"{op['Оплачено']:,.2f}"
        ws[f'H{row}'] = f"{op['Долг']:,.2f}"
        ws[f'I{row}'] = "0.00"
        for col in range(2, 10):
            ws.cell(row=row, column=col).alignment = center
            ws.cell(row=row, column=col).border = border
        row += 1

    # Обороты за период
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = "Обороты за период"
    ws[f'D{row}'] = f"{sum(op['Сумма'] for op in act_data):,.2f}"
    ws[f'E{row}'] = f"{sum(op['Оплачено'] for op in act_data):,.2f}"
    ws[f'F{row}'] = f"{sum(op['Долг'] for op in act_data):,.2f}"
    ws[f'G{row}'] = f"{sum(op['Оплачено'] for op in act_data):,.2f}"
    ws[f'H{row}'] = f"{sum(op['Долг'] for op in act_data):,.2f}"
    ws[f'I{row}'] = f"{sum(op['Сумма'] for op in act_data):,.2f}"
    for col in range(2, 10):
        ws.cell(row=row, column=col).alignment = center
        ws.cell(row=row, column=col).font = bold
        ws.cell(row=row, column=col).border = border
    row += 1

    # Сальдо конечное
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = "Сальдо конечное"
    ws[f'D{row}'] = f"{saldo_end:,.2f}"
    ws[f'E{row}'] = "0.00"
    ws[f'F{row}'] = "0.00"
    ws[f'G{row}'] = "0.00"
    ws[f'H{row}'] = "0.00"
    ws[f'I{row}'] = f"{saldo_end:,.2f}"
    for col in range(2, 10):
        ws.cell(row=row, column=col).alignment = center
        ws.cell(row=row, column=col).font = bold
        ws.cell(row=row, column=col).border = border
    row += 2

    # Итоговая строка с суммой прописью
    sum_words = num2words(abs(saldo_end), lang='ru').capitalize()
    if saldo_end < 0:
        sum_words = f"минус {sum_words}"
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'] = f"В пользу {company2} {saldo_end:,.2f} сум ({sum_words} сум)"
    ws[f'B{row}'].font = bold
    row += 2

    # Подписи
    ws[f'B{row}'] = f"От {company1}"
    ws[f'G{row}'] = f"От {company2}"
    row += 2
    ws[f'B{row}'] = "Директор"
    ws[f'G{row}'] = "Директор"
    row += 2
    ws[f'B{row}'] = "М.П."
    ws[f'G{row}'] = "М.П."

    # Ширина колонок
    widths = [5, 18, 40, 15, 15, 15, 15, 15, 15]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Сохраняем файл
    filename = f"reconciliation_act_{company1}_{company2}_{period_start}_{period_end}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# Пример использования
if __name__ == "__main__":
    # Тестовые данные
    test_data = [
        ('AVTOLIDER', 11113, 'Подфарник боковой', datetime(2024, 12, 10, 22, 10, 21), 
         'Продажа', 12.0, 35000.0, 420000.0, 'Ожидает оплаты'),
        ('AVTOLIDER', 10611, 'Лист рессорная задняя  25*90  Nº2  173см—Д', 
         datetime(2024, 12, 10, 22, 10, 21), 'Продажа', 1.0, 500000.0, 500000.0, 'Ожидает оплаты'),
        ('AVTOLIDER', 10939, 'Втулка торсиона кабины HOWO фторопласт LEO', 
         datetime(2024, 12, 10, 22, 10, 21), 'Продажа', 2.0, 30000.0, 60000.0, 'Ожидает оплаты')
    ]
    
    # Генерация накладной
    invoice_path = generate_invoice_excel(test_data)
    print(f"Накладная создана: {invoice_path}")
